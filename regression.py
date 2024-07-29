from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from scipy.optimize import minimize

def read_excel_to_matrix(file_path):
    #Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    #Initialize an empty matrix
    data_X = []
    data_y = []

    #Iterate through the rows in the sheet
    for X in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=X, column=col).value
            row_data.append(float(cell_value))
        data_X.append(row_data)
    
    for y in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=y, column=7).value
        data_y.append(float(cell_value))

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    #Close the workbook
    workbook.close()
    
    return array_X, array_y

def trainData(X, y):
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

    model = LinearRegression()
    model.fit(X_train, y_train)

    # Make predictions
    y_pred = model.predict(X_test)

    # Evaluate the model
    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)

    print(f'Mean Squared Error: {mse}')
    print(f'R-squared: {r2}')

    coefficients = model.coef_
    print("Model Coefficients:")
    for i, coef in enumerate(coefficients):
        print(f'Coefficient for feature {i+1}: {coef}')

    return model, y_test, y_pred


def plot_regression_results(y_test, y_pred):
    # Plot Actual vs Predicted values
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test, y_pred, edgecolor='k', alpha=0.7)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], 'k--', lw=2)
    plt.xlabel('Actual')
    plt.ylabel('Predicted')
    plt.title('Actual vs Predicted Values')
    plt.show()

    # Plot Residuals
    residuals = y_test - y_pred
    plt.figure(figsize=(10, 6))
    plt.scatter(y_pred, residuals, edgecolor='k', alpha=0.7)
    plt.axhline(y=0, color='r', linestyle='--')
    plt.xlabel('Predicted')
    plt.ylabel('Residuals')
    plt.title('Residuals Plot')
    plt.show()

def analyze_target_variable(y):
    # Convert the target array to a DataFrame for easier analysis
    df = pd.DataFrame(y, columns=['target'])

    # Check for missing values
    missing_values = df.isnull().sum()
    print("Missing Values:\n", missing_values)

    # Summary statistics
    summary_stats = df.describe()
    print("\nSummary Statistics:\n", summary_stats)

    # Visualize the distribution of the target variable
    plt.figure(figsize=(10, 6))
    sns.histplot(df['target'], kde=True)
    plt.title('Target Variable Distribution')
    plt.xlabel('Target')
    plt.ylabel('Frequency')
    plt.show()

    # Box plot to identify outliers
    plt.figure(figsize=(10, 6))
    sns.boxplot(x=df['target'])
    plt.title('Target Variable Box Plot')
    plt.xlabel('Target')
    plt.show()

    # Scatter plot to visualize relationships and detect outliers
    plt.figure(figsize=(10, 6))
    sns.scatterplot(x=df.index, y=df['target'])
    plt.title('Target Variable Scatter Plot')
    plt.xlabel('Index')
    plt.ylabel('Target')
    plt.show()

def objective_function(X, model):
    return model.predict(X.reshape(1, -1))[0]

def generate_lowest_target_with_constraints(model, feature_shape):
    # Initial guess for the features (mean of the dataset)
    initial_guess = np.zeros(feature_shape)

    # Define bounds for the optimization
    bounds = [(None, None)] * feature_shape  # No bounds

    # Define nonlinear constraints
    constraints = [
        {'type': 'eq', 'fun': constraint1},
        {'type': 'eq', 'fun': constraint2}
    ]

    result = minimize(objective_function, initial_guess, args=(model,), method='SLSQP', 
                      bounds=bounds, constraints=constraints)

    if result.success:
        optimized_features = result.x
        print(f"Optimized Features for Minimum Target: {optimized_features}")
        print(f"Predicted Minimum Target Value: {result.fun}")
        return optimized_features
    else:
        print("Optimization failed details:")
        print("Status:", result.status)
        print("Message:", result.message)
        raise ValueError("Optimization failed") 

def constraint1(X):
    return np.sin(np.radians(X[4])) * X[3] - np.sin(np.radians(X[2])) * X[1] - 20.73

def constraint2(X):
    return np.cos(np.radians(X[2])) * X[1] + np.cos(np.radians(X[4])) * X[3] + X[0] - 3001.20  # Note: X[5] should be X[0] in Python 0-based index

#Define the file path
file_path = 'dataset.xlsx'

#Read the data into a matrix
X , y = read_excel_to_matrix(file_path)

print("Feature Matrix X:\n")
print(X)

print("Target Array y:\n")
print(y)

analyze_target_variable(y)
model, y_test, y_pred = trainData(X, y)

#Generate new five parameters (X) for the lowest value (y) possible
optimized_features = generate_lowest_target_with_constraints(model, X.shape[1])

#Plot regression results
plot_regression_results(y_test, y_pred)
