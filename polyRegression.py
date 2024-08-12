import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression

# Function to read data from an Excel file into feature matrix X and target array y
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_X = []
    data_y = []

    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        data_X.append(row_data)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=sheet.max_column).value
        data_y.append(float(cell_value))  # Convert each value to float

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    return array_X, array_y

# Define the polynomial regression model fitting and evaluation
def polyRegression(X, y):
    # Split the dataset into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Generate polynomial features
    poly = PolynomialFeatures(degree=2)
    X_poly_train = poly.fit_transform(X_train)
    X_poly_test = poly.transform(X_test)

    # Fit the polynomial regression model
    poly_model = LinearRegression()
    poly_model.fit(X_poly_train, y_train)

    # Predict and calculate MSE and R-squared
    poly_pred = poly_model.predict(X_poly_test)
    poly_mse = mean_squared_error(y_test, poly_pred)
    poly_r2 = r2_score(y_test, poly_pred)
    
    print(f"Polynomial Regression Model MSE: {poly_mse}")
    print(f"Polynomial Regression Model R-squared: {poly_r2}")

    # Return the fitted model and polynomial transformer
    return poly_model, poly

# Updated function to read independent variables for prediction from specific columns (2nd to 6th) and rows (starting from 2nd)
def read_excel_for_prediction(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    new_data = []
    for row in range(2, sheet.max_row + 1):  # Start from row 2
        row_data = []
        for col in range(2, 7):  # Columns from 2 to 6 (inclusive)
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        new_data.append(row_data)

    array_new_data = np.array(new_data)
    workbook.close()
    return array_new_data

# Function to write predictions to the 8th row of the Excel file
def write_predictions_to_excel(file_path, predictions):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Print predictions and write them to the 8th row
    for i, prediction in enumerate(predictions):
        print(f"Prediction for data row {i + 1}: {prediction}")
        sheet.cell(row=i + 2, column=8, value=prediction)  # Write prediction in the 8th row

    workbook.save(file_path)
    workbook.close()

def predict_new(poly_model, poly, new_data):
    # Transform the new data using the polynomial features
    new_data_poly = poly.transform(new_data)  # Pass new_data directly, it's already 2D
    
    # Predict the dependent variable using the model
    predictions = poly_model.predict(new_data_poly)
    
    return predictions  # Return the predictions as an array

# Define the file path
file_path = 'dataset.xlsx'
prediction_file_path = 'prediction_data.xlsx'

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

print("Feature Matrix X:\n", X)
print("Target Array y:\n", y)

# Fit the polynomial regression model
poly_model, poly = polyRegression(X, y)

# Read the new data for prediction from another Excel file
new_data = read_excel_for_prediction(prediction_file_path)

# Predict for the new data points and write to Excel
predictions = predict_new(poly_model, poly, new_data)
write_predictions_to_excel(prediction_file_path, predictions)

print("Predictions have been written to the 8th row of the Excel file.")

