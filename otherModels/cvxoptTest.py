import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from cvxopt import matrix, solvers

# Function to read data from an Excel file into feature matrix X and target array y
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_X = []
    data_y = []

    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        data_X.append(row_data)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=sheet.max_column).value
        data_y.append(float(cell_value))  # Convert each value to float

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    return array_X, array_y

# Define the linear regression model fitting and evaluation
def linearRegression(X, y):
    # Split the dataset into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Fit the linear regression model
    lin_model = LinearRegression()
    lin_model.fit(X_train, y_train)

    # Print feature importances (coefficients)
    feature_importances = lin_model.coef_
    print("Feature Importances (Coefficients):")
    for i, coef in enumerate(feature_importances):
        print(f"Feature {i+1}: {coef}")

    return lin_model

# Normalize the data
def normalize_data(X):
    mean = np.mean(X, axis=0)
    std = np.std(X, axis=0)
    X_normalized = (X - mean) / std
    return X_normalized, mean, std

# Convert the problem to a cvxopt quadratic programming problem
def solve_with_cvxopt(X, y, initial_guess):
    n_features = X.shape[1]
    
    # Normalize the data
    X, mean, std = normalize_data(X)
    initial_guess_normalized = (initial_guess - mean) / std

    # Construct the P (Hessian matrix) for the quadratic term in the objective function
    P = 2 * matrix(np.dot(X.T, X))
    
    # Construct the q vector (linear term in the objective function)
    q = matrix(-2 * np.dot(X.T, y))
    
    # Set bounds to keep variables in a reasonable range
    lower_bounds = np.array([-3] * n_features)  # Example lower bound
    upper_bounds = np.array([3] * n_features)   # Example upper bound
    G = matrix(np.vstack([-np.eye(n_features), np.eye(n_features)]))
    h = matrix(np.hstack([-lower_bounds, upper_bounds]), (2 * n_features, 1))  # Convert h to a column vector

    # Constructing the equality constraint matrix A and vector b
    # A * X_optimized = b
    A = np.zeros((3, 5))  # 3 constraints, 5 variables
    b = np.zeros(3)

    # Constraint 1: np.sin(np.radians(X[4])) * X[1] - 575.75 = 0
    A[0, 1] = np.sin(np.radians(initial_guess_normalized[4]))
    b[0] = 575.75

    # Constraint 2: np.sin(np.radians(X[3])) * X[0] - 690.33 = 0
    A[1, 0] = np.sin(np.radians(initial_guess_normalized[3]))
    b[1] = 690.33

    # Constraint 3: np.cos(np.radians(X[3])) * X[0] + np.cos(np.radians(X[4])) * X[1] + X[2] - 3001.2 = 0
    A[2, 0] = np.cos(np.radians(initial_guess_normalized[3]))
    A[2, 1] = np.cos(np.radians(initial_guess_normalized[4]))
    A[2, 2] = 1.0
    b[2] = 3001.2

    # Convert A and b to matrix form for cvxopt
    A = matrix(A)
    b = matrix(b)

    # Solve the QP problem
    solution = solvers.qp(P, q, G, h, A, b)
    
    # Extract the optimized feature values
    X_optimized_normalized = np.array(solution['x']).flatten()

    # Denormalize the result
    X_optimized = X_optimized_normalized * std + mean
    
    return X_optimized

# Define the file path
file_path = 'dataset.xlsx'

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

print("Feature Matrix X:\n", X)
print("Target Array y:\n", y)

# Fit the linear regression model
lin_model = linearRegression(X, y)

# Initial guess for the optimization
initial_guess = np.array([706.46, 677.49, 2493.97, 77.73, 58.19])

# Solve the optimization problem using cvxopt
X_optimized = solve_with_cvxopt(X, y, initial_guess)

print("Optimized Feature Values (X):", X_optimized)

# Use the optimized features to predict the target y
y_predicted = lin_model.predict([X_optimized])

print("Predicted Minimum Target Value (y):", y_predicted[0])
