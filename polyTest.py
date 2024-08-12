import numpy as np
from sklearn.model_selection import train_test_split, cross_val_score, KFold, learning_curve
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.stats as stats
from scipy.stats import zscore

# Function to read data from an Excel file into feature matrix X and target array y
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_X = []
    data_y = []

    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        data_X.append(row_data)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=sheet.max_column).value
        data_y.append(float(cell_value))  # Convert each value to float

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    return array_X, array_y

# Define the polynomial regression model fitting and evaluation with overfitting check
def polyRegression(X, y):
    # Split the dataset into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Generate polynomial features
    poly = PolynomialFeatures(degree=2)
    X_poly_train = poly.fit_transform(X_train)
    X_poly_test = poly.transform(X_test)

    # Fit the polynomial regression model
    poly_model = LinearRegression()
    poly_model.fit(X_poly_train, y_train)

    # Predict and calculate MSE and R-squared for training set
    train_pred = poly_model.predict(X_poly_train)
    train_mse = mean_squared_error(y_train, train_pred)
    train_r2 = r2_score(y_train, train_pred)
    
    print(f"Polynomial Regression Model Training MSE: {train_mse}")
    print(f"Polynomial Regression Model Training R-squared: {train_r2}")

    # Predict and calculate MSE and R-squared for test set
    test_pred = poly_model.predict(X_poly_test)
    test_mse = mean_squared_error(y_test, test_pred)
    test_r2 = r2_score(y_test, test_pred)
    
    print(f"Polynomial Regression Model Test MSE: {test_mse}")
    print(f"Polynomial Regression Model Test R-squared: {test_r2}")

    # Check for overfitting
    if train_mse < test_mse and (test_r2 < train_r2):
        print("The model may be overfitting. Consider using a simpler model or applying regularization.")

from sklearn.model_selection import cross_val_score, KFold

def cross_validation_analysis(X, y, degree=2):
    # Generate polynomial features
    poly = PolynomialFeatures(degree=degree)
    X_poly = poly.fit_transform(X)

    # Define the model
    poly_model = LinearRegression()

    # Define the k-fold cross-validation
    kf = KFold(n_splits=5, shuffle=True, random_state=42)
    
    # Perform cross-validation
    mse_scores = cross_val_score(poly_model, X_poly, y, scoring='neg_mean_squared_error', cv=kf)
    r2_scores = cross_val_score(poly_model, X_poly, y, scoring='r2', cv=kf)
    
    # Convert negative MSE scores to positive
    mse_scores = -mse_scores
    
    print(f"Cross-Validation MSE Scores: {mse_scores}")
    print(f"Cross-Validation R-squared Scores: {r2_scores}")
    print(f"Mean MSE: {np.mean(mse_scores)}, Standard Deviation: {np.std(mse_scores)}")
    print(f"Mean R-squared: {np.mean(r2_scores)}, Standard Deviation: {np.std(r2_scores)}")
    
    return mse_scores, r2_scores

def plot_learning_curves(X, y, degree=2):
    poly = PolynomialFeatures(degree=degree)
    X_poly = poly.fit_transform(X)
    poly_model = LinearRegression()

    train_sizes, train_scores, test_scores = learning_curve(
        poly_model, X_poly, y, cv=5, scoring='neg_mean_squared_error', train_sizes=np.linspace(0.1, 1.0, 10), random_state=42)

    train_scores_mean = -train_scores.mean(axis=1)
    test_scores_mean = -test_scores.mean(axis=1)

    plt.figure(figsize=(10, 6))
    plt.plot(train_sizes, train_scores_mean, 'o-', color="r", label="Training MSE")
    plt.plot(train_sizes, test_scores_mean, 'o-', color="g", label="Cross-validation MSE")
    plt.xlabel("Training Size")
    plt.ylabel("MSE")
    plt.title("Learning Curves")
    plt.legend(loc="best")
    plt.grid()
    plt.show()

def residual_analysis(X, y, degree=2):
    poly = PolynomialFeatures(degree=degree)
    X_poly = poly.fit_transform(X)
    
    # Fit the model
    poly_model = LinearRegression()
    poly_model.fit(X_poly, y)
    
    # Predict the target
    y_pred = poly_model.predict(X_poly)
    
    # Calculate residuals
    residuals = y - y_pred
    
    # Plot residuals
    plt.figure(figsize=(10, 6))
    sns.residplot(x=y_pred, y=residuals, lowess=True, line_kws={'color': 'red', 'lw': 1})
    plt.xlabel("Predicted Values")
    plt.ylabel("Residuals")
    plt.title("Residual Plot")
    plt.grid()
    plt.show()
    
    # Plot histogram of residuals
    plt.figure(figsize=(10, 6))
    sns.histplot(residuals, kde=True)
    plt.title("Residuals Distribution")
    plt.grid()
    plt.show()
    
    # Q-Q plot for normality
    plt.figure(figsize=(10, 6))
    stats.probplot(residuals, dist="norm", plot=plt)
    plt.title("Q-Q Plot")
    plt.grid()
    plt.show()

# Define the file path
file_path = 'dataset.xlsx'

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

# Fit the polynomial regression model and check for overfitting
polyRegression(X, y)

# Run cross-validation analysis
cross_validation_analysis(X, y)

# Plot learning curves
plot_learning_curves(X, y)

# Perform residual analysis
residual_analysis(X, y)
