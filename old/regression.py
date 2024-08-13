import numpy as np
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score
from scipy.optimize import minimize
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import make_pipeline

# Function to read data from an Excel file into feature matrix X and target array y
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_X = []
    data_y = []

    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        data_X.append(row_data)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=sheet.max_column).value
        data_y.append(float(cell_value))  # Convert each value to float

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    return array_X, array_y

# Function to preprocess data (convert to numeric and handle missing values)
def preprocess_data(X, y):
    X = pd.DataFrame(X)
    y = pd.Series(y)

    X = X.apply(pd.to_numeric, errors='coerce').fillna(0).values
    y = pd.to_numeric(y, errors='coerce').fillna(0).values

    return X, y

# Function to train a linear regression model and evaluate it
def train_model(X, y):
    X, y = preprocess_data(X, y)
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    model = LinearRegression()
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)

    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)

    print(f'Mean Squared Error: {mse}')
    print(f'R-squared: {r2}')

    coefficients = model.coef_
    for i, coef in enumerate(coefficients):
        print(f'Coefficient for feature {i+1}: {coef}')

    return model, y_test, y_pred

# Function to plot regression results (Actual vs Predicted and Residuals)
def plot_regression_results(y_test, y_pred):
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test, y_pred, edgecolor='k', alpha=0.7)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], 'k--', lw=2)
    plt.xlabel('Actual')
    plt.ylabel('Predicted')
    plt.title('Actual vs Predicted Values')
    plt.show()

    residuals = y_test - y_pred
    plt.figure(figsize=(10, 6))
    plt.scatter(y_pred, residuals, edgecolor='k', alpha=0.7)
    plt.axhline(y=0, color='r', linestyle='--')
    plt.xlabel('Predicted')
    plt.ylabel('Residuals')
    plt.title('Residuals Plot')
    plt.show()

# Function to analyze the target variable (distribution, summary statistics, and outliers)
def analyze_target_variable(y):
    df = pd.DataFrame(y, columns=['target'])

    missing_values = df.isnull().sum()
    print("Missing Values:\n", missing_values)

    summary_stats = df.describe()
    print("\nSummary Statistics:\n", summary_stats)

    plt.figure(figsize=(10, 6))
    sns.histplot(df['target'], kde=True)
    plt.title('Target Variable Distribution')
    plt.xlabel('Target')
    plt.ylabel('Frequency')
    plt.show()

    plt.figure(figsize=(10, 6))
    sns.boxplot(x=df['target'])
    plt.title('Target Variable Box Plot')
    plt.xlabel('Target')
    plt.show()

    plt.figure(figsize=(10, 6))
    sns.scatterplot(x=df.index, y=df['target'])
    plt.title('Target Variable Scatter Plot')
    plt.xlabel('Index')
    plt.ylabel('Target')
    plt.show()

# Objective function for optimization (predicting target using the model)
def objective_function(X, model):
    return model.predict(X.reshape(1, -1))[0]

# Function to generate optimized features to achieve the lowest target value under constraints
def generate_lowest_target_with_constraints(model, X):
    feature_shape = X.shape[1]
    initial_guess = np.mean(X, axis=0)
    bounds = [(min(X[:, i]), max(X[:, i])) for i in range(feature_shape)]
    
    constraints = [
        {'type': 'eq', 'fun': lambda X: constraint1(X)},
        {'type': 'eq', 'fun': lambda X: constraint2(X)},
        {'type': 'eq', 'fun': lambda X: constraint3(X)}
    ]
    
    result = minimize(lambda x: objective_function(x, model), initial_guess, method='SLSQP', 
                      bounds=bounds, constraints=constraints)

    if result.success:
        optimized_features = result.x
        print(f"Optimized Features for Minimum Target: {optimized_features}")
        print(f"Predicted Minimum Target Value: {result.fun}")
        return optimized_features
    else:
        print("Optimization failed details:")
        print("Status:", result.status)
        print("Message:", result.message)
        raise ValueError("Optimization failed")

# Constraint functions for the optimization problem
def constraint1(X):
    return np.sin(np.radians(X[4])) * X[1] - 575.75

def constraint2(X):
    return np.sin(np.radians(X[3])) * X[0] - 690.33

def constraint3(X):
    return np.cos(np.radians(X[3])) * X[0] + np.cos(np.radians(X[4])) * X[1] + X[2] - 3001.2

# Define the file path
file_path = 'dataset.xlsx'

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

print("Feature Matrix X:\n", X)
print("Target Array y:\n", y)

# Analyze the target variable
analyze_target_variable(y)

# Train the model and evaluate
model, y_test, y_pred = train_model(X, y)

# Generate new parameters for the lowest possible target value
optimized_features = generate_lowest_target_with_constraints(model, X)

# Plot regression results
plot_regression_results(y_test, y_pred)
