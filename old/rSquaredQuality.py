import pandas as pd
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score

# Function to train the model and evaluate R-squared quality
def evaluate_model_quality(X, y):
    # Split the data into training and test sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

    # Train the model
    model = LinearRegression()
    model.fit(X_train, y_train)

    # Make predictions
    y_train_pred = model.predict(X_train)
    y_test_pred = model.predict(X_test)

    # Calculate R-squared for training and test sets
    r2_train = r2_score(y_train, y_train_pred)
    r2_test = r2_score(y_test, y_test_pred)
    
    print(f'R-squared on training set: {r2_train}')
    print(f'R-squared on test set: {r2_test}')

    # Residual analysis
    residuals = y_test - y_test_pred
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test_pred, residuals, edgecolor='k', alpha=0.7)
    plt.axhline(y=0, color='r', linestyle='--')
    plt.xlabel('Predicted')
    plt.ylabel('Residuals')
    plt.title('Residuals Plot')
    plt.show()

    # Calculate Adjusted R-squared
    n = len(y_test)  # Number of observations
    k = X.shape[1]   # Number of predictors

    def adjusted_r2(r2, n, k):
        return 1 - ((1 - r2) * (n - 1)) / (n - k - 1)

    r2_adj = adjusted_r2(r2_test, n, k)
    print(f'Adjusted R-squared: {r2_adj}')

    # Cross-validation
    cv_scores = cross_val_score(model, X, y, cv=5, scoring='r2')
    print(f'Cross-validated R-squared scores: {cv_scores}')
    print(f'Mean Cross-validated R-squared: {np.mean(cv_scores)}')

    return model, r2_train, r2_test, r2_adj, cv_scores

# Example usage with data loading and processing
# Define the file path
file_path = 'dataset.xlsx'

# Function to read the data into a matrix
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    data_X = []
    data_y = []

    for X in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=X, column=col).value
            row_data.append(cell_value)
        data_X.append(row_data)
    
    for y in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=y, column=sheet.max_column).value
        data_y.append(cell_value)

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    
    return array_X, array_y

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

# Convert data to DataFrame/Series for proper conversion and handling
X = pd.DataFrame(X)
y = pd.Series(y)

# Convert data to numeric and handle missing values
X = X.apply(pd.to_numeric, errors='coerce').fillna(0).values
y = pd.to_numeric(y, errors='coerce').fillna(0).values

# Evaluate model quality
model, r2_train, r2_test, r2_adj, cv_scores = evaluate_model_quality(X, y)
