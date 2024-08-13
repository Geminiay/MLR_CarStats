import numpy as np
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error, r2_score
from scipy.optimize import differential_evolution, NonlinearConstraint

# Function to read data from an Excel file into feature matrix X and target array y
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_X = []
    data_y = []

    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        data_X.append(row_data)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=sheet.max_column).value
        data_y.append(float(cell_value))  # Convert each value to float

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    return array_X, array_y

# Function to preprocess data (convert to numeric and handle missing values)
def preprocess_data(X, y):
    X = pd.DataFrame(X)
    y = pd.Series(y)

    X = X.apply(pd.to_numeric, errors='coerce').fillna(0).values
    y = pd.to_numeric(y, errors='coerce').fillna(0).values

    return X, y

# Function to train a random forest regression model and evaluate it
def train_model(X, y):
    X, y = preprocess_data(X, y)
    
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    model = RandomForestRegressor(random_state=42)
    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)

    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)

    print(f'Mean Squared Error: {mse}')
    print(f'R-squared: {r2}')

    importances = model.feature_importances_
    for i, importance in enumerate(importances):
        print(f'Importance for feature {i+1}: {importance}')

    return model, X_test, y_test, y_pred

# Function to plot regression results (Actual vs Predicted and Residuals)
def plot_regression_results(y_test, y_pred):
    plt.figure(figsize=(10, 6))
    plt.scatter(y_test, y_pred, edgecolor='k', alpha=0.7)
    plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], 'k--', lw=2)
    plt.xlabel('Actual')
    plt.ylabel('Predicted')
    plt.title('Actual vs Predicted Values')
    plt.show()

    residuals = y_test - y_pred
    plt.figure(figsize=(10, 6))
    plt.scatter(y_pred, residuals, edgecolor='k', alpha=0.7)
    plt.axhline(y=0, color='r', linestyle='--')
    plt.xlabel('Predicted')
    plt.ylabel('Residuals')
    plt.title('Residuals Plot')
    plt.show()

# Function to analyze the target variable (distribution, summary statistics, and outliers)
def analyze_target_variable(y):
    df = pd.DataFrame(y, columns=['target'])

    missing_values = df.isnull().sum()
    print("Missing Values:\n", missing_values)

    summary_stats = df.describe()
    print("\nSummary Statistics:\n", summary_stats)

    plt.figure(figsize=(10, 6))
    sns.histplot(df['target'], kde=True)
    plt.title('Target Variable Distribution')
    plt.xlabel('Target')
    plt.ylabel('Frequency')
    plt.show()

    plt.figure(figsize=(10, 6))
    sns.boxplot(x=df['target'])
    plt.title('Target Variable Box Plot')
    plt.xlabel('Target')
    plt.show()

    plt.figure(figsize=(10, 6))
    sns.scatterplot(x=df.index, y=df['target'])
    plt.title('Target Variable Scatter Plot')
    plt.xlabel('Index')
    plt.ylabel('Target')
    plt.show()

# Custom objective function for grid search
def objective_function(features, model):
    return model.predict([features])[0]

# Constraints for the optimization
def constraint_1(x):
    return np.sin(np.radians(x[4])) * x[1] - 575.75

def constraint_2(x):
    return np.sin(np.radians(x[3])) * x[0] - 690.33

def constraint_3(x):
    return np.cos(np.radians(x[3])) * x[0] + np.cos(np.radians(x[4])) * x[1] + x[2] - 3001.2

# Wrapper function for the optimization to include constraints
def constrained_objective_function(features, model):
    return objective_function(features, model)

# Function to generate optimized features using Differential Evolution
def generate_lowest_target_with_differential_evolution(model, X):
    feature_shape = X.shape[1]
    
    # Define the bounds for each feature
    bounds = [(np.min(X[:, i]), np.max(X[:, i])) for i in range(feature_shape)]
    
    # Define the constraints as NonlinearConstraint objects
    constraints = [
        NonlinearConstraint(constraint_1, 0, 0),
        NonlinearConstraint(constraint_2, 0, 0),
        NonlinearConstraint(constraint_3, 0, 0)
    ]
    
    result = differential_evolution(
        constrained_objective_function,
        bounds,
        args=(model,),
        constraints=constraints,
        strategy='best1bin',
        maxiter=5000,
        popsize=15,
        tol=1e-8,
        mutation=(0.5, 1),
        recombination=0.7
    )
    
    best_features = result.x
    min_target_value = result.fun
    
    print(f"Optimized Features for Minimum Target: {best_features}")
    print(f"Predicted Minimum Target Value: {min_target_value}")
    return best_features

# Function to verify if constraints are satisfied
def verify_constraints(features):
    c1 = np.sin(np.radians(features[4])) * features[1] - 575.75
    c2 = np.sin(np.radians(features[3])) * features[0] - 690.33
    c3 = np.cos(np.radians(features[3])) * features[0] + np.cos(np.radians(features[4])) * features[1] + features[2] - 3001.2
    
    return c1, c2, c3

# Define the file path
file_path = 'dataset.xlsx'

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

print("Feature Matrix X:\n", X)
print("Target Array y:\n", y)

# Analyze the target variable
analyze_target_variable(y)

# Train the model and evaluate
model, X_test, y_test, y_pred = train_model(X, y)

# Generate new parameters for the lowest possible target value using Differential Evolution
optimized_features = generate_lowest_target_with_differential_evolution(model, X)

# Verify constraints
constraints = verify_constraints(optimized_features)
print("Constraint 1 (should be 0):", constraints[0])
print("Constraint 2 (should be 0):", constraints[1])
print("Constraint 3 (should be 0):", constraints[2])

# Plot regression results
plot_regression_results(y_test, y_pred)
