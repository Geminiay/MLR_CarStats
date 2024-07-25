import openpyxl
import random
from openpyxl.styles import Alignment

def CdMaker(filename):
    
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    for row in range(3, sheet.max_row + 1):
        sheet.cell(row=row, column=7).value = round(random.uniform(0,1),3)
        sheet.cell(row=row, column=7).alignment = Alignment(horizontal='center')

    workbook.save(filename)
    print("The random values has been added.")

filename = input("Enter the file name to generate random Cd's:\n") + ".xlsx"

CdMaker(filename)
