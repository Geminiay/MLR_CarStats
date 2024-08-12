import openpyxl
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score
from sklearn.preprocessing import PolynomialFeatures
from sklearn.linear_model import LinearRegression
from scipy.optimize import minimize

# Function to read data from an Excel file into feature matrix X and target array y
def read_excel_to_matrix(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data_X = []
    data_y = []

    for row in range(2, sheet.max_row + 1):
        row_data = []
        for col in range(2, sheet.max_column):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(float(cell_value))  # Convert each value to float
        data_X.append(row_data)

    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=sheet.max_column).value
        data_y.append(float(cell_value))  # Convert each value to float

    array_X = np.array(data_X)
    array_y = np.array(data_y)

    workbook.close()
    return array_X, array_y

# Define the polynomial regression model fitting and evaluation
def polyRegression(X, y):
    # Split the dataset into training and testing sets
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

    # Generate polynomial features
    poly = PolynomialFeatures(degree=2)
    X_poly_train = poly.fit_transform(X_train)
    X_poly_test = poly.transform(X_test)

    # Fit the polynomial regression model
    poly_model = LinearRegression()
    poly_model.fit(X_poly_train, y_train)

    # Predict and calculate MSE and R-squared
    poly_pred = poly_model.predict(X_poly_test)
    poly_mse = mean_squared_error(y_test, poly_pred)
    poly_r2 = r2_score(y_test, poly_pred)
    
    print(f"Polynomial Regression Model MSE: {poly_mse}")
    print(f"Polynomial Regression Model R-squared: {poly_r2}")

    # Return the fitted model and polynomial transformer
    return poly_model, poly

# Constraint functions for the optimization problem
def constraint1(X):
    return np.sin(np.radians(X[4])) * X[1] - 575.75

def constraint2(X):
    return np.sin(np.radians(X[3])) * X[0] - 690.33

def constraint3(X):
    return np.cos(np.radians(X[3])) * X[0] + np.cos(np.radians(X[4])) * X[1] + X[2] - 3001.2

# Optimization objective function
def objective_function(X_input, poly_model, poly):
    # Reshape input data to match the expected shape
    X_input = np.array(X_input).reshape(1, -1)
    
    # Transform the input data to polynomial features
    X_poly_input = poly.transform(X_input)
    
    # Predict the target value using the model
    y_pred = poly_model.predict(X_poly_input)
    
    # Return the predicted value (we want to minimize this)
    return y_pred[0]


# Define the file path
file_path = 'dataset.xlsx'

# Read the data into a matrix
X, y = read_excel_to_matrix(file_path)

print("Feature Matrix X:\n", X)
print("Target Array y:\n", y)

# Fit the polynomial regression model
poly_model, poly = polyRegression(X, y)

# Initial guess for the optimizer (could be the mean of the input data)
initial_guess = np.mean(X, axis=0)

# Bounds for the optimizer (you can set specific bounds for your problem if needed)
bounds = [(min(X[:, i]), max(X[:, i])) for i in range(X.shape[1])]

# Define constraints in a format suitable for the minimize function
constraints = [
    {'type': 'ineq', 'fun': constraint1},
    {'type': 'ineq', 'fun': constraint2},
    {'type': 'ineq', 'fun': constraint3}
]

# Use a minimization algorithm to find the feature values that minimize the target value
result = minimize(objective_function, initial_guess, args=(poly_model, poly), bounds=bounds, constraints=constraints, method='SLSQP')

# Extract the optimized feature values
X_optimized = result.x
y_optimized = result.fun

print("Optimized Feature Values (X):", X_optimized)
print("Predicted Minimum Target Value (y):", y_optimized)

